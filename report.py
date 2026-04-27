import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_report(rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Позиции Kaspi"

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Отчёт позиций Kaspi — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Код товара", "Название", "Позиция (#)", "Страница", "Место на стр.", "Проверено"]
    col_widths = [15, 45, 14, 12, 15, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[2].height = 22

    for row_idx, row in enumerate(rows, start=3):
        is_even = (row_idx % 2 == 0)
        bg = PatternFill("solid", fgColor="F0F4FF" if is_even else "FFFFFF")

        checked_at = row["checked_at"] or ""
        try:
            dt = datetime.strptime(checked_at[:19], "%Y-%m-%d %H:%M:%S")
            checked_str = dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            checked_str = checked_at

        position = row["position"]
        pos_str = f"#{position}" if position else "не найден"

        values = [
            row["code"],
            row["product_name"] or "—",
            pos_str,
            row["page"] or "—",
            row["place_on_page"] or "—",
            checked_str,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = bg
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 3 and position:
                cell.font = Font(bold=True, color="1A73E8")
            if col_idx == 1:
                cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
